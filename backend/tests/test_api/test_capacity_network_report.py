
import pytest
from fastapi.testclient import TestClient
from sqlalchemy.orm import Session
from app.models.capacity_network import CapacityNetworkValues
from app.models.capacity_network import CapacityNetworkValues, ZoneDeviceMappingNetwork, RegionZoneMappingNetwork
import io
import openpyxl

def test_get_dashboard_empty(client: TestClient, normal_user_token_headers):
    """Test getting dashboard data with no data in DB"""
    response = client.get(
        "/api/v1/capacity-network-report/dashboard",
        params={"region": "XYZ", "production_hours": True},
        headers=normal_user_token_headers
    )
    assert response.status_code == 200
    data = response.json()
    assert data["region"] == "XYZ"
    assert data["zone_summary"] == []

def test_dashboard_with_data(client: TestClient, test_db: Session, normal_user_token_headers):
    """Test dashboard with seeded data"""
    # Seed data
    region = RegionZoneMappingNetwork(region_name="XYZ", zone_name="Zone A")
    test_db.add(region)
    
    mapping = ZoneDeviceMappingNetwork(zone_name="Zone A", device_name="Device 1")
    test_db.add(mapping)
    
    cap = CapacityNetworkValues(
        device_name="Device 1",
        mean_cpu=50.0, peak_cpu=80.0,
        mean_memory=40.0, peak_memory=60.0,
        cpu_time="10:00", memory_time="10:00"
    )
    test_db.add(cap)
    test_db.commit()

    response = client.get(
        "/api/v1/capacity-network-report/dashboard",
        params={"region": "XYZ", "production_hours": True},
        headers=normal_user_token_headers
    )
    assert response.status_code == 200
    data = response.json()
    summary = data["zone_summary"][0]
    assert summary["zone_name"] == "Zone A"
    assert summary["total_device_count"] == 1
    # 80% peak cpu > 70 implies Critical (Red)
    assert summary["cpu_critical"] == 1
    assert summary["cpu_normal"] == 0

def test_add_device_zone_mapping(client: TestClient, test_db: Session, normal_user_token_headers):
    """Test adding a device to a zone"""
    payload = {
        "zone_name": "New Zone",
        "device_name": "New Device"
    }
    response = client.post(
        "/api/v1/capacity-network-report/device-zone-mapping/add",
        json=payload,
        headers=normal_user_token_headers
    )
    assert response.status_code == 201
    assert response.json()["message"] == "Device 'New Device' successfully added to zone 'New Zone'"

    # Verify in DB
    mapping = test_db.query(ZoneDeviceMappingNetwork).filter_by(device_name="New Device").first()
    assert mapping is not None
    assert mapping.zone_name == "New Zone"

def test_delete_device_zone_mapping(client: TestClient, test_db: Session, normal_user_token_headers):
    """Test deleting a device-zone mapping"""
    # Seed
    mapping = ZoneDeviceMappingNetwork(zone_name="Delete Zone", device_name="Delete Device")
    test_db.add(mapping)
    test_db.commit()

    response = client.delete(
        "/api/v1/capacity-network-report/device-zone-mapping/delete",
        params={"zone_name": "Delete Zone", "device_name": "Delete Device"},
        headers=normal_user_token_headers
    )
    assert response.status_code == 200
    
    # Verify gone
    mapping = test_db.query(ZoneDeviceMappingNetwork).filter_by(device_name="Delete Device").first()
    assert mapping is None

def test_add_zone_region_mapping(client: TestClient, test_db: Session, normal_user_token_headers):
    """Test adding a zone to a region"""
    payload = {
        "region_name": "XYZ",
        "zone_name": "New Zone Region"
    }
    response = client.post(
        "/api/v1/capacity-network-report/zone-region-mapping/add",
        json=payload,
        headers=normal_user_token_headers
    )
    assert response.status_code == 201
    
    # Verify in DB
    mapping = test_db.query(RegionZoneMappingNetwork).filter_by(zone_name="New Zone Region").first()
    assert mapping is not None
    assert mapping.region_name == "XYZ"

def test_update_zone_region_mapping(client: TestClient, test_db: Session, normal_user_token_headers):
    """Test updating a zone name"""
    # Seed
    mapping = RegionZoneMappingNetwork(region_name="XYZ", zone_name="Old Name")
    test_db.add(mapping)
    test_db.commit()

    response = client.put(
        "/api/v1/capacity-network-report/zone-region-mapping/update",
        json={"region_name": "XYZ", "zone_name": "Old Name"},
        params={"new_zone_name": "New Name"},
        headers=normal_user_token_headers
    )
    assert response.status_code == 200

    # Verify
    updated = test_db.query(RegionZoneMappingNetwork).filter_by(zone_name="New Name").first()
    assert updated is not None
    assert updated.region_name == "XYZ"

def test_export_devices(client: TestClient, test_db: Session, normal_user_token_headers):
    """Test exporting devices to Excel"""
    # Seed data
    test_db.add(RegionZoneMappingNetwork(region_name="XYZ", zone_name="XYZ Zone A"))
    test_db.add(ZoneDeviceMappingNetwork(zone_name="XYZ Zone A", device_name="Device 1"))
    test_db.add(CapacityNetworkValues(device_name="Device 1", mean_cpu=50.0, peak_cpu=80.0))
    test_db.commit()

    response = client.get(
        "/api/v1/capacity-network-report/export",
        headers=normal_user_token_headers
    )
    assert response.status_code == 200
    assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    
    # Verify excel content
    f = io.BytesIO(response.content)
    wb = openpyxl.load_workbook(f)
    assert "XYZ" in wb.sheetnames
    ws = wb["XYZ"]
    # Check header
    assert ws.cell(row=1, column=1).value == "Zone Name"
    # Check data
    # Row 2 should be our data
    assert ws.cell(row=2, column=1).value == "XYZ Zone A"
    assert ws.cell(row=2, column=2).value == "Device 1"

def test_export_summary(client: TestClient, test_db: Session, normal_user_token_headers):
    """Test exporting summary to Excel"""
    # Seed data
    test_db.add(RegionZoneMappingNetwork(region_name="XYZ", zone_name="XYZ Zone A"))
    test_db.add(ZoneDeviceMappingNetwork(zone_name="XYZ Zone A", device_name="Device 1"))
    test_db.commit()

    response = client.get(
        "/api/v1/capacity-network-report/export-summary",
        headers=normal_user_token_headers
    )
    assert response.status_code == 200
    assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    
    f = io.BytesIO(response.content)
    wb = openpyxl.load_workbook(f)
    assert "Production XYZ" in wb.sheetnames
