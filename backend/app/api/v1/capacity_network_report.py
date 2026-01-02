from fastapi import APIRouter, Depends, File, UploadFile, Query, HTTPException, status
from fastapi.responses import StreamingResponse
from sqlalchemy import func, or_, and_, text
from sqlalchemy.orm import Session
from typing import Dict, List, Tuple, Optional
from pydantic import BaseModel
import io
import openpyxl
import pandas as pd
import re
from datetime import datetime
import tempfile
import os
from pathlib import Path

from app.api.v1.auth import get_current_active_user
from app.core.database import get_db, get_engine
from app.core.logging_config import get_logger
from app.models.user import User
from app.models.capacity_network import CapacityNetworkValues
from app.models.capacity_network import ZoneDeviceMappingNetwork, RegionZoneMappingNetwork

router = APIRouter(prefix="/capacity-network-report", tags=["capacity-network-report"])
logger = get_logger(__name__)

# Storage directory for capacity network report files
BACKEND_DIR = Path(__file__).parent.parent.parent.parent
CAPACITY_NETWORK_UPLOADS_DIR = BACKEND_DIR / "uploads" / "capacity_network_reports"
CAPACITY_NETWORK_UPLOADS_DIR.mkdir(parents=True, exist_ok=True)


def convert_NaN_to_None(value):
    """Function to convert NaN values to None so blank values can be inserted into database"""
    return None if pd.isna(value) else value


def changeFormat(value):
    """Convert value to float, truncating to integer if >= 1.0"""
    try:
        if value is None:
            return None
        value = float(value)
        return value if value < 1.0 else float(int(value))
    except (ValueError, TypeError):
        return None


def extract_date_time_peak(dt_string):
    """Function to extract date, time and peak value from the string"""
    if not isinstance(dt_string, str):
        return None
    pattern = r"(\d{2}-\w{3}-\d{4}) (\d{2}:\d{2}) \d+\.\d+\(peak: (\d+\.\d+)\) to \d{2}-\w{3}-\d{4} (\d{2}:\d{2})"
    match = re.search(pattern, dt_string)
    if match:
        date, start_time, peak_value, end_time = match.groups()
        return date, start_time, peak_value, end_time
    else:
        logger.warning(f"No match found for string: {dt_string}")
        return None


def is_between(cur_date, start_date, end_date):
    """Check if cur_date is between start_date and end_date (inclusive)"""
    try:
        cur_date_obj = datetime.strptime(cur_date, "%d-%b-%Y")
        start_date_obj = datetime.strptime(start_date, "%Y-%m-%d")
        end_date_obj = datetime.strptime(end_date, "%Y-%m-%d")
        return start_date_obj <= cur_date_obj <= end_date_obj
    except (ValueError, TypeError) as e:
        logger.error(f"Error parsing dates: {e}")
        return False


# ============================================================================
# Database Interaction Functions (Currently Commented Out)
# These functions will be used when Excel files are available
# ============================================================================

def delete_Capacity_Network_Values_table(db: Session) -> dict:
    """
    Delete all records from Capacity_Network_Values database using ORM.
    """
    try:
        db.query(CapacityNetworkValues).delete()
        db.commit()
        return {
            "status": "success",
            "message": "All rows deleted for capacity_network_values table",
        }
    except Exception as e:
        db.rollback()
        logger.error(f"Error deleting capacity network values: {str(e)}")
        raise


def insert_capacity_network_values(
    db: Session,
    device_name: str,
    mean_cpu: Optional[float],
    peak_cpu: Optional[float],
    ntimes_cpu: int,
    mean_memory: Optional[float],
    peak_memory: Optional[float],
    ntimes_memory: int,
) -> dict:
    """
    Insert a new CapacityNetworkValues record into the database.
    """
    try:
        new_entry = CapacityNetworkValues(
            device_name=device_name,
            mean_cpu=mean_cpu,
            peak_cpu=peak_cpu,
            ntimes_cpu=ntimes_cpu,
            mean_memory=mean_memory,
            peak_memory=peak_memory,
            ntimes_memory=ntimes_memory,
        )
        db.add(new_entry)
        db.commit()
        return {
            "status": "success",
            "message": "Capacity Network Values data inserted successfully",
        }
    except Exception as e:
        db.rollback()
        logger.error(f"Error inserting capacity network values: {str(e)}")
        raise


def update_ntime_cpu_network(db: Session, device_name: str) -> dict:
    """
    Increment ntimes_cpu counter for a device.
    """
    try:
        record = db.query(CapacityNetworkValues).filter_by(device_name=device_name).first()
        if record:
            record.ntimes_cpu = (record.ntimes_cpu or 0) + 1
            db.commit()
        return {
            "status": "success",
            "message": "ntimes cpu value increased",
        }
    except Exception as e:
        db.rollback()
        logger.error(f"Error updating ntime_cpu: {str(e)}")
        raise


def update_capacity_peaks_cpu_network(
    db: Session,
    device_name: str,
    cpu_peak: Optional[float],
    cpu_date: Optional[str],
    cpu_time: Optional[str],
    cpu_alert_duration: Optional[float],
) -> dict:
    """
    Update CPU peak related values for a device.
    """
    try:
        record = db.query(CapacityNetworkValues).filter_by(device_name=device_name).first()
        if record:
            record.peak_cpu = cpu_peak
            record.cpu_date = cpu_date
            record.cpu_time = cpu_time
            record.cpu_alert_duration = cpu_alert_duration
            db.commit()
        return {
            "status": "success",
            "message": "cpu peak related values updated",
        }
    except Exception as e:
        db.rollback()
        logger.error(f"Error updating capacity peaks cpu: {str(e)}")
        raise


def update_ntime_memory_network(db: Session, device_name: str) -> dict:
    """
    Increment ntimes_memory counter for a device.
    """
    try:
        record = db.query(CapacityNetworkValues).filter_by(device_name=device_name).first()
        if record:
            record.ntimes_memory = (record.ntimes_memory or 0) + 1
            db.commit()
        return {
            "status": "success",
            "message": "ntimes memory value increased",
        }
    except Exception as e:
        db.rollback()
        logger.error(f"Error updating ntime_memory: {str(e)}")
        raise


def update_capacity_peaks_memory_network(
    db: Session,
    device_name: str,
    memory_peak: Optional[float],
    memory_date: Optional[str],
    memory_time: Optional[str],
    memory_alert_duration: Optional[float],
) -> dict:
    """
    Update Memory peak related values for a device.
    """
    try:
        record = db.query(CapacityNetworkValues).filter_by(device_name=device_name).first()
        if record:
            record.peak_memory = memory_peak
            record.memory_date = memory_date
            record.memory_time = memory_time
            record.memory_alert_duration = memory_alert_duration
            db.commit()
        return {
            "status": "success",
            "message": "memory peak related values updated",
        }
    except Exception as e:
        db.rollback()
        logger.error(f"Error updating capacity peaks memory: {str(e)}")
        raise


async def process_capacity_network_files(
    db: Session,
    network_data_file_path: str,
) -> None:
    """
    Process network capacity file and store data into capacity_network_values table.
    Reads Excel file from stored file path, extracts capacity metrics, and updates the database.
    
    NOTE: Currently disabled - function does nothing as Excel files are not available yet.
    All database operations are commented out.
    """
    # Logic commented out as per request
    return True

    # logger.info(f"process_capacity_network_files called with file: {network_data_file_path}")
    # logger.info(f"Starting capacity network file processing")
    # 
    # try:
    #     # Read Excel file
    #     # Adjust cols based on network report format (assuming similar to capacity report for now)
    #     df_input1 = pd.read_excel(network_data_file_path, skiprows=3, usecols="I,J,K,L,M,N,O,R,T")
    #     
    #     # Re-define the columns
    #     df_input1.columns = [
    #         "device_name",
    #         "mean_cpu",
    #         "peak_cpu",
    #         "ntimes_cpu",
    #         "mean_memory",
    #         "peak_memory",
    #         "ntimes_memory",
    #         "first_peak_cpu_dt",
    #         "first_peak_memory_dt"
    #     ]
    #     
    #     # Delete previous Capacity Network values table data
    #     delete_Capacity_Network_Values_table(db)
    #     logger.info("Deleting previous capacity network values")
    #     
    #     # Last device name for the alerts which corresponds to the row with last device name
    #     last_ntimes_cpu = 0
    #     last_ntimes_memory = 0
    #     last_device = None
    #     cur_cpu_peak = None
    #     cur_memory_peak = None
    #     
    #     # Loop over the raw sheet to fill capacity values database
    #     for _, row in df_input1.iterrows():
    #         device_name = row["device_name"]
    #         
    #         # Skip empty rows or rows with device_name == "1"
    #         if row.isnull().all() or device_name == "1":
    #             continue
    #         
    #         ntimes_cpu = row.iloc[3]
    #         ntimes_memory = row.iloc[6]
    #         
    #         # Update last device tracking
    #         if pd.notna(device_name):
    #             last_ntimes_cpu = ntimes_cpu
    #             last_ntimes_memory = ntimes_memory
    #             last_device = device_name
    #             cur_cpu_peak = None
    #             cur_memory_peak = None
    #         
    #         # Handle peak values when ntimes are 0
    #         if ntimes_cpu == 0 and ntimes_memory == 0:
    #             cur_cpu_peak = convert_NaN_to_None(row.get("peak_cpu"))
    #             cur_memory_peak = convert_NaN_to_None(row.get("peak_memory"))
    #         elif ntimes_memory == 0:
    #             cur_memory_peak = convert_NaN_to_None(row.get("peak_memory"))
    #         elif ntimes_cpu == 0:
    #             cur_cpu_peak = convert_NaN_to_None(row.get("peak_cpu"))
    #         
    #         # Insert initial capacity values if both ntimes are 0
    #         if ntimes_cpu == 0 and ntimes_memory == 0:
    #             insert_capacity_network_values(
    #                 db,
    #                 device_name,
    #                 changeFormat(convert_NaN_to_None(row.get("mean_cpu"))),
    #                 changeFormat(cur_cpu_peak),
    #                 0,
    #                 changeFormat(convert_NaN_to_None(row.get("mean_memory"))),
    #                 changeFormat(cur_memory_peak),
    #                 0
    #             )
    #             continue
    #         
    #         # Current ntimes values
    #         cur_ntimes_cpu = last_ntimes_cpu
    #         cur_ntimes_memory = last_ntimes_memory
    #         
    #         # Process CPU details
    #         if cur_ntimes_cpu > 0:
    #             cpu_details = extract_date_time_peak(row.get("first_peak_cpu_dt", ""))
    #             # Note: logic for checking dates against start/end would go here if we had date inputs
    #             if cpu_details and cpu_details[2] is not None:
    #             
    #                 update_ntime_cpu_network(db, last_device)
    #                 
    #                 # Update CPU peak if higher
    #                 if cur_cpu_peak is None or float(cpu_details[2]) > cur_cpu_peak:
    #                     cur_cpu_peak = float(cpu_details[2])
    #                     cpu_date = cpu_details[0]
    #                     cpu_time = cpu_details[1]
    #                     cpu_alert_duration = (
    #                         datetime.strptime(cpu_details[3], "%H:%M")
    #                         - datetime.strptime(cpu_details[1], "%H:%M")
    #                     ).total_seconds() / 60
    #                 
    #                     update_capacity_peaks_cpu_network(
    #                         db, 
    #                         last_device, 
    #                         cur_cpu_peak,
    #                         cpu_date,
    #                         cpu_time,
    #                         cpu_alert_duration
    #                     )
    #             
    #         # Process Memory details
    #         if cur_ntimes_memory > 0:
    #             memory_details = extract_date_time_peak(row.get("first_peak_memory_dt", ""))
    #             if memory_details and memory_details[2] is not None:
    #             
    #                 update_ntime_memory_network(db, last_device)
    #                 
    #                 # Update Memory peak if higher
    #                 if cur_memory_peak is None or float(memory_details[2]) > cur_memory_peak:
    #                     cur_memory_peak = float(memory_details[2])
    #                     memory_date = memory_details[0]
    #                     memory_time = memory_details[1]
    #                     memory_alert_duration = (
    #                         datetime.strptime(memory_details[3], "%H:%M")
    #                         - datetime.strptime(memory_details[1], "%H:%M")
    #                     ).total_seconds() / 60
    #                 
    #                     update_capacity_peaks_memory_network(
    #                         db,
    #                         last_device,
    #                         cur_memory_peak,
    #                         memory_date,
    #                         memory_time,
    #                         memory_alert_duration
    #                     )
    #         
    #         # Update last ntimes values
    #         last_ntimes_cpu = cur_ntimes_cpu
    #         last_ntimes_memory = cur_ntimes_memory
    #     
    #     logger.info("Capacity network file processing completed successfully")
    # 
    # except Exception as e:
    #     logger.error(f"Error processing capacity network files: {str(e)}", exc_info=True)
    #     db.rollback()
    #     raise HTTPException(
    #         status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
    #         detail=f"Failed to process capacity network files: {str(e)}"
    #     )

@router.post("/upload")
async def upload_capacity_network_report(
    network_data_file: UploadFile = File(...),
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    """
    Accept network capacity data: single file upload.
    File is saved to persistent storage for later processing.
    """
    try:
        # Create timestamp-based directory for this upload
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        upload_dir = CAPACITY_NETWORK_UPLOADS_DIR / timestamp
        upload_dir.mkdir(parents=True, exist_ok=True)
        
        # Save network data file
        network_data_path = upload_dir / f"network_data_{network_data_file.filename or 'file.xlsx'}"
        network_data_content = await network_data_file.read()
        with open(network_data_path, "wb") as f:
            f.write(network_data_content)
        logger.info(f"Saved network data file to: {network_data_path}")
        
        # Process network capacity file using stored file path
        await process_capacity_network_files(
            db=db,
            network_data_file_path=str(network_data_path),
        )

        return {
            "message": "Network capacity report file uploaded and saved successfully.",
            "network_data_filename": network_data_file.filename,
            "upload_directory": str(upload_dir),
            "network_data_path": str(network_data_path),
        }
    except Exception as e:
        logger.error(f"Error uploading network capacity file: {str(e)}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to upload network capacity file: {str(e)}"
        )


@router.get("/export")
async def export_capacity_network_devices(
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    """
    Export all network capacity data organized by region into an Excel file.
    Each region gets its own sheet with device details.
    """
    
    def get_data_for_region(region_prefix: str) -> List[Tuple]:
        """Query devices for a given region prefix."""
        results = (
            db.query(
                ZoneDeviceMappingNetwork.zone_name,
                ZoneDeviceMappingNetwork.device_name,
                CapacityNetworkValues.mean_cpu,
                CapacityNetworkValues.peak_cpu,
                CapacityNetworkValues.mean_memory,
                CapacityNetworkValues.peak_memory,
            )
            .join(
                CapacityNetworkValues,
                ZoneDeviceMappingNetwork.device_name == CapacityNetworkValues.device_name,
                isouter=True,
            )
            .filter(ZoneDeviceMappingNetwork.zone_name.startswith(region_prefix))
            .order_by(
                ZoneDeviceMappingNetwork.zone_name,
                ZoneDeviceMappingNetwork.device_name,
            )
            .all()
        )
        return results

    wb = openpyxl.Workbook()
    
    sheets_created = False

    regions = ["XYZ", "ORM-XYZ", "DRM", "ORM-DRM"]
    headers = [
        "Zone Name",
        "Device Name",
        "Mean CPU",
        "Peak CPU",
        "Mean Memory",
        "Peak Memory",
    ]

    for region in regions:
        data = get_data_for_region(region)
        if data:
            ws = wb.create_sheet(title=region)
            ws.append(headers)
            for row in data:
                ws.append(list(row))

            # Auto-adjust column widths
            for col_cells in ws.columns:
                length = max(
                    len(str(cell.value)) if cell.value is not None else 0
                    for cell in col_cells
                )
                ws.column_dimensions[col_cells[0].column_letter].width = min(
                    max(length + 2, 12), 40
                )
            sheets_created = True

    if sheets_created:
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
    else:
        ws = wb.active
        ws.title = "No Data"
        ws.append(["No devices found"])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="capacity_network_devices.xlsx"'
        },
    )


@router.get("/export-summary")
async def export_capacity_network_summary(
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    """
    Generate network capacity summary Excel with separate sheets per region
    and production/non-production hours.
    
    Sheet layout (per region):
        Zone Name
        Total Device Count
        CPU Green
        Memory Green
        CPU Yellow
        Memory Yellow
        CPU Red
        Memory Red
    """

    def get_zones_for_region(region_prefix: str) -> List[str]:
        """Return distinct zone names for a region prefix (e.g. 'XYZ')."""
        results: List[Tuple[str]] = (
            db.query(ZoneDeviceMappingNetwork.zone_name)
            .filter(ZoneDeviceMappingNetwork.zone_name.startswith(region_prefix))
            .distinct()
            .order_by(ZoneDeviceMappingNetwork.zone_name)
            .all()
        )
        return [row[0] for row in results]

    def compute_category(
        zone_name: str,
        prod_hours: bool,
        peak_min: int,
        peak_max: int,
        attribute,
        time_column,
    ) -> int:
        """
        Count distinct devices for a zone that fall into a threshold bucket
        for the given attribute (CPU/Memory) and time-of-day window.
        """
        # Production: 09:00-16:00, Non-production: everything else
        time_filter = (
            time_column.between("09:00", "16:00")
            if prod_hours
            else ~time_column.between("09:00", "16:00")
        )

        query = (
            db.query(func.count(CapacityNetworkValues.device_name.distinct()))
            .join(
                ZoneDeviceMappingNetwork,
                CapacityNetworkValues.device_name == ZoneDeviceMappingNetwork.device_name,
            )
            .filter(
                ZoneDeviceMappingNetwork.zone_name == zone_name,
                attribute.between(peak_min, peak_max),
                time_filter,
            )
        )

        count = query.scalar()
        return int(count or 0)

    def build_region_summary(region_prefix: str, prod_hours: bool) -> List[List]:
        """
        Build summary rows for a region and production flag.

        Returns list of:
            [zone, total_devices, cpu_green, memory_green,
             cpu_yellow, memory_yellow, cpu_red, memory_red]
        """
        zones = get_zones_for_region(region_prefix)
        summary_rows: List[List] = []

        for zone in zones:
            # Total devices (no time filter)
            total_devices = (
                db.query(func.count(ZoneDeviceMappingNetwork.device_name.distinct()))
                .filter(ZoneDeviceMappingNetwork.zone_name == zone)
                .scalar()
            )
            total_devices = int(total_devices or 0)

            # CPU categories
            cpu_red = compute_category(
                zone,
                prod_hours,
                71,
                100,
                CapacityNetworkValues.peak_cpu,
                CapacityNetworkValues.cpu_time,
            )
            cpu_yellow = compute_category(
                zone,
                prod_hours,
                61,
                70,
                CapacityNetworkValues.peak_cpu,
                CapacityNetworkValues.cpu_time,
            )
            cpu_green = max(total_devices - (cpu_red + cpu_yellow), 0)

            # Memory categories
            memory_red = compute_category(
                zone,
                prod_hours,
                71,
                100,
                CapacityNetworkValues.peak_memory,
                CapacityNetworkValues.memory_time,
            )
            memory_yellow = compute_category(
                zone,
                prod_hours,
                61,
                70,
                CapacityNetworkValues.peak_memory,
                CapacityNetworkValues.memory_time,
            )
            memory_green = max(total_devices - (memory_red + memory_yellow), 0)

            summary_rows.append(
                [
                    zone,
                    total_devices,
                    cpu_green,
                    memory_green,
                    cpu_yellow,
                    memory_yellow,
                    cpu_red,
                    memory_red,
                ]
            )

        # Sort zones alphabetically
        summary_rows.sort(key=lambda r: r[0] or "")
        return summary_rows

    headers = [
        "Zone Name",
        "Total Device Count",
        "CPU Green",
        "Memory Green",
        "CPU Yellow",
        "Memory Yellow",
        "CPU Red",
        "Memory Red",
    ]

    # Create workbook with separate sheets for each region and prod flag
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    regions = ["XYZ", "ORM-XYZ", "DRM", "ORM-DRM"]

    for region in regions:
        # Production sheet
        prod_rows = build_region_summary(region, prod_hours=True)
        if prod_rows:
            ws_prod = wb.create_sheet(title=f"Production {region}")
            ws_prod.append(headers)
            for row in prod_rows:
                ws_prod.append(row)

            for col_cells in ws_prod.columns:
                length = max(
                    len(str(cell.value)) if cell.value is not None else 0
                    for cell in col_cells
                )
                ws_prod.column_dimensions[col_cells[0].column_letter].width = min(
                    max(length + 2, 12), 40
                )

        # Non-production sheet
        non_prod_rows = build_region_summary(region, prod_hours=False)
        if non_prod_rows:
            ws_non = wb.create_sheet(title=f"Non-Production {region}")
            ws_non.append(headers)
            for row in non_prod_rows:
                ws_non.append(row)

            for col_cells in ws_non.columns:
                length = max(
                    len(str(cell.value)) if cell.value is not None else 0
                    for cell in col_cells
                )
                ws_non.column_dimensions[col_cells[0].column_letter].width = min(
                    max(length + 2, 12), 40
                )

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="capacity_network_summary.xlsx"'
        },
    )

# Pydantic models for requests
class DeviceZoneMappingNetworkRequest(BaseModel):
    zone_name: str
    device_name: str

class ZoneRegionMappingNetworkRequest(BaseModel):
    zone_name: str
    region_name: str


@router.get("/dashboard")
async def get_capacity_network_dashboard(
    region: str = Query(..., description="Region name: XYZ, ORM-XYZ, DRM, or ORM-DRM"),
    production_hours: bool = Query(True, description="True for production hours (09:00-16:00), False for non-production"),
    zone_name: Optional[str] = Query(None, description="Optional: Get device details for a specific zone"),
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    """
    Get network capacity dashboard data for a region.
    Returns zone summary and optionally device details for a specific zone.
    """
    # Validate region
    valid_regions = ["XYZ", "ORM-XYZ", "DRM", "ORM-DRM"]
    if region not in valid_regions:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Invalid region. Must be one of: {', '.join(valid_regions)}"
        )

    # Get zones for the region from region_zone_mapping_network
    zones_query = (
        db.query(RegionZoneMappingNetwork.zone_name)
        .filter(RegionZoneMappingNetwork.region_name == region)
        .order_by(RegionZoneMappingNetwork.zone_name)
        .all()
    )
    zones = [z[0] for z in zones_query]

    def compute_category_count(
        zone_name: str,
        prod_hours: bool,
        peak_min: float,
        peak_max: float,
        attribute,
        time_column,
    ) -> int:
        """Count devices in a category (Normal/Warning/Critical) for a zone."""
        # Production: 09:00-16:00, Non-production: everything else
        if prod_hours:
            time_filter = time_column.between("09:00", "16:00")
        else:
            time_filter = or_(
                time_column < "09:00",
                time_column > "16:00",
                time_column.is_(None)
            )

        query = (
            db.query(CapacityNetworkValues.device_name)
            .join(
                ZoneDeviceMappingNetwork,
                CapacityNetworkValues.device_name == ZoneDeviceMappingNetwork.device_name,
            )
            .filter(
                ZoneDeviceMappingNetwork.zone_name == zone_name,
                attribute.between(peak_min, peak_max),
                time_filter,
            )
        )
        
        count = query.distinct().count()
        return int(count or 0)

    # Build zone summary
    zone_summary = []
    for zone in zones:
        # Total device count (no time filter)
        total_devices = (
            db.query(ZoneDeviceMappingNetwork.device_name)
            .filter(ZoneDeviceMappingNetwork.zone_name == zone)
            .distinct()
            .count()
        )
        total_devices = int(total_devices or 0)

        # CPU categories
        cpu_critical = compute_category_count(zone, production_hours, 71, 100, CapacityNetworkValues.peak_cpu, CapacityNetworkValues.cpu_time)
        cpu_warning = compute_category_count(zone, production_hours, 61, 70, CapacityNetworkValues.peak_cpu, CapacityNetworkValues.cpu_time)
        cpu_normal = max(total_devices - (cpu_critical + cpu_warning), 0)

        # Memory categories
        memory_critical = compute_category_count(zone, production_hours, 71, 100, CapacityNetworkValues.peak_memory, CapacityNetworkValues.memory_time)
        memory_warning = compute_category_count(zone, production_hours, 61, 70, CapacityNetworkValues.peak_memory, CapacityNetworkValues.memory_time)
        memory_normal = max(total_devices - (memory_critical + memory_warning), 0)

        zone_summary.append({
            "zone_name": zone,
            "total_device_count": total_devices,
            "cpu_normal": cpu_normal,
            "cpu_warning": cpu_warning,
            "cpu_critical": cpu_critical,
            "memory_normal": memory_normal,
            "memory_warning": memory_warning,
            "memory_critical": memory_critical,
        })

    response = {
        "region": region,
        "production_hours": production_hours,
        "zone_summary": zone_summary,
    }

    # If zone_name is provided, also return device details for that zone
    if zone_name:
        # Verify zone belongs to the region
        zone_exists = (
            db.query(RegionZoneMappingNetwork)
            .filter(
                RegionZoneMappingNetwork.region_name == region,
                RegionZoneMappingNetwork.zone_name == zone_name
            )
            .first()
        )
        
        if not zone_exists:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Zone '{zone_name}' not found in region '{region}'"
            )

        # Get devices for this zone
        devices_query = (
            db.query(
                CapacityNetworkValues.device_name,
                CapacityNetworkValues.peak_cpu,
                CapacityNetworkValues.ntimes_cpu,
                CapacityNetworkValues.cpu_alert_duration,
                CapacityNetworkValues.peak_memory,
                CapacityNetworkValues.ntimes_memory,
                CapacityNetworkValues.memory_alert_duration,
            )
            .join(
                ZoneDeviceMappingNetwork,
                CapacityNetworkValues.device_name == ZoneDeviceMappingNetwork.device_name,
            )
            .filter(ZoneDeviceMappingNetwork.zone_name == zone_name)
            .order_by(CapacityNetworkValues.device_name)
            .all()
        )

        device_details = []
        for device in devices_query:
            device_details.append({
                "device_name": device[0],
                "cpu_peak_percent": device[1],
                "cpu_peak_count": device[2],
                "cpu_peak_duration_min": device[3],
                "memory_peak_percent": device[4],
                "memory_peak_count": device[5],
                "memory_peak_duration_min": device[6],
            })

        response["device_details"] = device_details

    return response


@router.get("/zones")
async def get_all_zones_network(
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    """Get all zones from region_zone_mapping_network"""
    zones = (
        db.query(RegionZoneMappingNetwork.zone_name, RegionZoneMappingNetwork.region_name)
        .order_by(RegionZoneMappingNetwork.region_name, RegionZoneMappingNetwork.zone_name)
        .all()
    )
    return {
        "zones": [{"zone_name": z[0], "region_name": z[1]} for z in zones]
    }


@router.get("/devices")
async def get_all_devices_network(
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    """Get all unique device names from capacity_network_values"""
    devices = (
        db.query(CapacityNetworkValues.device_name)
        .filter(CapacityNetworkValues.device_name.isnot(None))
        .distinct()
        .order_by(CapacityNetworkValues.device_name)
        .all()
    )
    return {
        "devices": [d[0] for d in devices if d[0]]
    }


@router.get("/regions")
async def get_all_regions_network(
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    """Get all unique regions from region_zone_mapping_network"""
    regions = (
        db.query(RegionZoneMappingNetwork.region_name)
        .distinct()
        .order_by(RegionZoneMappingNetwork.region_name)
        .all()
    )
    return {
        "regions": [r[0] for r in regions]
    }


@router.get("/zone-device-mappings")
async def get_zone_device_mappings_network(
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    """Get all zone-device mappings"""
    mappings = (
        db.query(ZoneDeviceMappingNetwork.zone_name, ZoneDeviceMappingNetwork.device_name)
        .order_by(ZoneDeviceMappingNetwork.zone_name, ZoneDeviceMappingNetwork.device_name)
        .all()
    )
    return {
        "mappings": [{"zone_name": m[0], "device_name": m[1]} for m in mappings]
    }


@router.post("/device-zone-mapping/add", status_code=status.HTTP_201_CREATED)
async def add_device_to_zone_network(
    request: DeviceZoneMappingNetworkRequest,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    """Add a device to a zone (create zone-device mapping)"""
    logger.info(f"add_device_to_zone_network called with zone_name='{request.zone_name}', device_name='{request.device_name}'")
    
    # Check if device is already present in that zone
    existing = (
        db.query(ZoneDeviceMappingNetwork)
        .filter(
            ZoneDeviceMappingNetwork.zone_name == request.zone_name,
            ZoneDeviceMappingNetwork.device_name == request.device_name
        )
        .first()
    )
    
    if existing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"The device name '{request.device_name}' you are trying to add is already present in zone '{request.zone_name}'"
        )
    
    # Create device-zone mapping using ORM
    try:
        new_mapping = ZoneDeviceMappingNetwork(
            zone_name=request.zone_name,
            device_name=request.device_name
        )
        db.add(new_mapping)
        db.commit()
        
        return {
            "message": f"Device '{request.device_name}' successfully added to zone '{request.zone_name}'",
            "zone_name": request.zone_name,
            "device_name": request.device_name
        }
    except Exception as e:
        db.rollback()
        logger.error(f"Exception while adding device to zone: {str(e)}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to add device to zone: {str(e)}"
        )


@router.put("/device-zone-mapping/update")
async def update_device_zone_mapping_network(
    old_zone_name: str = Query(...),
    old_device_name: str = Query(...),
    new_zone_name: str = Query(None),
    new_device_name: str = Query(None),
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    """Update a device-zone mapping (move device to different zone or rename device)"""
    # Find existing mapping
    existing = (
        db.query(ZoneDeviceMappingNetwork)
        .filter(
            ZoneDeviceMappingNetwork.zone_name == old_zone_name,
            ZoneDeviceMappingNetwork.device_name == old_device_name
        )
        .first()
    )
    if not existing:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Mapping not found: device '{old_device_name}' in zone '{old_zone_name}'"
        )

    # Update zone if provided
    if new_zone_name and new_zone_name != old_zone_name:
        zone_exists = (
            db.query(RegionZoneMappingNetwork)
            .filter(RegionZoneMappingNetwork.zone_name == new_zone_name)
            .first()
        )
        if not zone_exists:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Zone '{new_zone_name}' not found"
            )
        existing.zone_name = new_zone_name

    # Update device if provided
    if new_device_name and new_device_name != old_device_name:
        existing.device_name = new_device_name

    try:
        db.commit()
        return {
            "message": "Device-zone mapping updated successfully",
            "zone_name": existing.zone_name,
            "device_name": existing.device_name
        }
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to update mapping: {str(e)}"
        )


@router.delete("/device-zone-mapping/delete")
async def delete_device_zone_mapping_network(
    zone_name: str = Query(...),
    device_name: str = Query(...),
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    """Delete a device-zone mapping"""
    mapping = (
        db.query(ZoneDeviceMappingNetwork)
        .filter(
            ZoneDeviceMappingNetwork.zone_name == zone_name,
            ZoneDeviceMappingNetwork.device_name == device_name
        )
        .first()
    )
    
    if not mapping:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Mapping not found: device '{device_name}' in zone '{zone_name}'"
        )
    
    try:
        db.delete(mapping)
        db.commit()
        return {
            "message": f"Device '{device_name}' removed from zone '{zone_name}' successfully"
        }
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to delete mapping: {str(e)}"
        )


@router.post("/zone-region-mapping/add", status_code=status.HTTP_201_CREATED)
async def add_zone_to_region_network(
    request: ZoneRegionMappingNetworkRequest,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    """Add a zone to a region (create zone-region mapping)"""
    existing = (
        db.query(RegionZoneMappingNetwork)
        .filter(
            RegionZoneMappingNetwork.region_name == request.region_name,
            RegionZoneMappingNetwork.zone_name == request.zone_name
        )
        .first()
    )
    
    if existing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Zone '{request.zone_name}' already exists in region '{request.region_name}'"
        )
    
    # Create zone-region mapping using ORM
    try:
        new_mapping = RegionZoneMappingNetwork(
            region_name=request.region_name,
            zone_name=request.zone_name
        )
        db.add(new_mapping)
        db.commit()
        
        return {
            "message": f"Zone '{request.zone_name}' successfully added to region '{request.region_name}'",
            "region_name": request.region_name,
            "zone_name": request.zone_name
        }
    except Exception as e:
        db.rollback()
        logger.error(f"Exception while adding zone to region: {str(e)}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to add zone to region: {str(e)}"
        )


@router.put("/zone-region-mapping/update")
async def update_zone_region_mapping_network(
    request: ZoneRegionMappingNetworkRequest,
    new_zone_name: str = Query(...),
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    """Update a zone name within a region"""
    existing = (
        db.query(RegionZoneMappingNetwork)
        .filter(
            RegionZoneMappingNetwork.region_name == request.region_name,
            RegionZoneMappingNetwork.zone_name == request.zone_name
        )
        .first()
    )
    
    if not existing:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Zone '{request.zone_name}' not found in region '{request.region_name}'"
        )
    
    # Update all zone_device_mapping_network entries
    db.query(ZoneDeviceMappingNetwork).filter(
        ZoneDeviceMappingNetwork.zone_name == request.zone_name
    ).update({"zone_name": new_zone_name})
    
    existing.zone_name = new_zone_name
    
    try:
        db.commit()
        return {
            "message": "Zone renamed successfully",
            "region_name": existing.region_name,
            "zone_name": existing.zone_name
        }
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to update zone: {str(e)}"
        )


@router.delete("/zone-region-mapping/delete")
async def delete_zone_region_mapping_network(
    zone_name: str = Query(...),
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    """Delete a zone (and all its device mappings)"""
    mapping = (
        db.query(RegionZoneMappingNetwork)
        .filter(RegionZoneMappingNetwork.zone_name == zone_name)
        .first()
    )
    
    if not mapping:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Zone '{zone_name}' not found"
        )
    
    try:
        # Delete all device mappings for this zone
        db.query(ZoneDeviceMappingNetwork).filter(
            ZoneDeviceMappingNetwork.zone_name == zone_name
        ).delete()
        
        # Delete the zone-region mapping
        db.delete(mapping)
        db.commit()
        
        return {
            "message": f"Zone '{zone_name}' and all its device mappings deleted successfully"
        }
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to delete zone: {str(e)}"
        )
