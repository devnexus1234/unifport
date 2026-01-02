from fastapi import APIRouter, Depends, File, Form, UploadFile, Query, HTTPException, status
from fastapi.responses import StreamingResponse
from sqlalchemy import func, or_, and_, text
from sqlalchemy.orm import Session
from typing import Dict, List, Tuple, Optional
from pydantic import BaseModel
import io
import openpyxl
import pandas as pd
import re
from datetime import datetime
import tempfile
import os
from pathlib import Path

from app.api.v1.auth import get_current_active_user
from app.core.database import get_db, get_engine
from app.core.logging_config import get_logger
from app.models.user import User
from app.models.capacity import CapacityValues
from app.models.capacity import ZoneDeviceMapping, RegionZoneMapping

router = APIRouter(prefix="/capacity-firewall-report", tags=["capacity-firewall-report"])
logger = get_logger(__name__)

# Storage directory for capacity report files
# Path is relative to the backend directory (where the app runs from)
# __file__ is backend/app/api/v1/capacity_report.py, so we go up 3 levels to backend/
BACKEND_DIR = Path(__file__).parent.parent.parent.parent
CAPACITY_UPLOADS_DIR = BACKEND_DIR / "uploads" / "capacity_reports"
CAPACITY_UPLOADS_DIR.mkdir(parents=True, exist_ok=True)


def convert_NaN_to_None(value):
    """Function to convert NaN values to None so blank values can be inserted into database"""
    return None if pd.isna(value) else value


def extract_date_time_peak(dt_string):
    """Function to extract date, time and peak value from the string"""
    if not isinstance(dt_string, str):
        return None
    pattern = r"(\d{2}-\w{3}-\d{4}) (\d{2}:\d{2}) \d+\.\d+\(peak: (\d+\.\d+)\) to \d{2}-\w{3}-\d{4} (\d{2}:\d{2})"
    match = re.search(pattern, dt_string)
    if match:
        date, start_time, peak_value, end_time = match.groups()
        return date, start_time, peak_value, end_time
    else:
        logger.warning(f"No match found for string: {dt_string}")
        return None


def is_between(cur_date, start_date, end_date):
    """Check if cur_date is between start_date and end_date (inclusive)"""
    try:
        cur_date_obj = datetime.strptime(cur_date, "%d-%b-%Y")
        start_date_obj = datetime.strptime(start_date, "%Y-%m-%d")
        end_date_obj = datetime.strptime(end_date, "%Y-%m-%d")
        return start_date_obj <= cur_date_obj <= end_date_obj
    except (ValueError, TypeError) as e:
        logger.error(f"Error parsing dates: {e}")
        return False


def changeFormat(value):
    """Convert value to float, truncating to integer if >= 1.0"""
    try:
        if value is None:
            return None
        value = float(value)
        return value if value < 1.0 else float(int(value))
    except (ValueError, TypeError):
        return None


# ============================================================================
# Database Interaction Functions (Currently Commented Out)
# These functions will be used when Excel files are available
# ============================================================================

def delete_Capacity_Values_table(db: Session) -> dict:
    """
    Delete all records from Capacity_values database using ORM.
    """
    try:
        db.query(CapacityValues).delete()
        db.commit()
        return {
            "status": "success",
            "message": "All rows deleted for capacity_values table",
        }
    except Exception as e:
        db.rollback()
        logger.error(f"Error deleting capacity values: {str(e)}")
        raise


def insert_capacity_values(
    db: Session,
    device_name: str,
    mean_cpu: Optional[float],
    peak_cpu: Optional[float],
    ntimes_cpu: int,
    mean_memory: Optional[float],
    peak_memory: Optional[float],
    ntimes_memory: int,
) -> dict:
    """
    Insert a new CapacityValues record into the database.
    """
    try:
        new_entry = CapacityValues(
            device_name=device_name,
            mean_cpu=mean_cpu,
            peak_cpu=peak_cpu,
            ntimes_cpu=ntimes_cpu,
            mean_memory=mean_memory,
            peak_memory=peak_memory,
            ntimes_memory=ntimes_memory,
        )
        db.add(new_entry)
        db.commit()
        return {
            "status": "success",
            "message": "Capacity Values data inserted successfully",
        }
    except Exception as e:
        db.rollback()
        logger.error(f"Error inserting capacity values: {str(e)}")
        raise


def update_ntime_cpu(db: Session, device_name: str) -> dict:
    """
    Increment ntimes_cpu counter for a device.
    """
    try:
        record = db.query(CapacityValues).filter_by(device_name=device_name).first()
        if record:
            record.ntimes_cpu = (record.ntimes_cpu or 0) + 1
            db.commit()
        return {
            "status": "success",
            "message": "ntimes cpu value increased",
        }
    except Exception as e:
        db.rollback()
        logger.error(f"Error updating ntime_cpu: {str(e)}")
        raise


def update_capacity_peaks_cpu(
    db: Session,
    device_name: str,
    cpu_peak: Optional[float],
    cpu_date: Optional[str],
    cpu_time: Optional[str],
    cpu_alert_duration: Optional[float],
) -> dict:
    """
    Update CPU peak related values for a device.
    """
    try:
        record = db.query(CapacityValues).filter_by(device_name=device_name).first()
        if record:
            record.peak_cpu = cpu_peak
            record.cpu_date = cpu_date
            record.cpu_time = cpu_time
            record.cpu_alert_duration = cpu_alert_duration
            db.commit()
        return {
            "status": "success",
            "message": "cpu peak related values updated",
        }
    except Exception as e:
        db.rollback()
        logger.error(f"Error updating capacity peaks cpu: {str(e)}")
        raise


def update_ntime_memory(db: Session, device_name: str) -> dict:
    """
    Increment ntimes_memory counter for a device.
    """
    try:
        record = db.query(CapacityValues).filter_by(device_name=device_name).first()
        if record:
            record.ntimes_memory = (record.ntimes_memory or 0) + 1
            db.commit()
        return {
            "status": "success",
            "message": "ntimes memory value increased",
        }
    except Exception as e:
        db.rollback()
        logger.error(f"Error updating ntime_memory: {str(e)}")
        raise


def update_capacity_peaks_memory(
    db: Session,
    device_name: str,
    memory_peak: Optional[float],
    memory_date: Optional[str],
    memory_time: Optional[str],
    memory_alert_duration: Optional[float],
) -> dict:
    """
    Update Memory peak related values for a device.
    """
    try:
        record = db.query(CapacityValues).filter_by(device_name=device_name).first()
        if record:
            record.peak_memory = memory_peak
            record.memory_date = memory_date
            record.memory_time = memory_time
            record.memory_alert_duration = memory_alert_duration
            db.commit()
        return {
            "status": "success",
            "message": "memory peak related values updated",
        }
    except Exception as e:
        db.rollback()
        logger.error(f"Error updating capacity peaks memory: {str(e)}")
        raise


def update_connection_in_capacity_values(
    db: Session,
    device_name: str,
    mean_connection: Optional[float],
    peak_connection: Optional[float],
) -> dict:
    """
    Update connection values (mean_connection and peak_connection) for a device.
    """
    try:
        record = db.query(CapacityValues).filter_by(device_name=device_name).first()
        if record:
            record.mean_connection = mean_connection
            record.peak_connection = peak_connection
            db.commit()
        return {
            "status": "success",
            "message": "Connection data updated successfully",
        }
    except Exception as e:
        db.rollback()
        logger.error(f"Error updating connection values: {str(e)}")
        raise


async def process_capacity_files(
    db: Session,
    start_date: str,
    end_date: str,
    raw_data_file_path: str,
    connection_count_file_path: str,
) -> None:
    """
    Process capacity files and store data into capacity_values table.
    Reads Excel files from stored file paths, extracts capacity metrics, and updates the database.
    
    NOTE: Currently disabled - function does nothing as Excel files are not available yet.
    All database operations are commented out.
    """
    # Logic commented out as per request
    return True

    # logger.info(f"process_capacity_files called with files: {raw_data_file_path}, {connection_count_file_path}")
    # 
    # raw_input_1 = raw_data_file_path
    # raw_input_2 = connection_count_file_path
    # 
    # try:
    #     # Read Excel files
    #     df_input1 = pd.read_excel(raw_input_1, skiprows=3, usecols="I,J,K,L,M,N,O,R,T")
    #     df_input2 = pd.read_excel(raw_input_2, skiprows=5, usecols="C,K,O")
    #     
    #     # Re-define the columns as in raw sheet, columns are nested, difficult to manipulate
    #     df_input1.columns = [
    #         "device_name",
    #         "mean_cpu",
    #         "peak_cpu",
    #         "ntimes_cpu",
    #         "mean_memory",
    #         "peak_memory",
    #         "ntimes_memory",
    #         "first_peak_cpu_dt",
    #         "first_peak_memory_dt"
    #     ]
    #     
    #     df_input2.columns = ["Entity Name", "Mean", "Peak"]
    #     
    #     # Delete previous Capacity values table data
    #     delete_Capacity_Values_table(db)
    #     
    #     # Last device name for the alerts which corresponds to the row with last device name
    #     last_ntimes_cpu = 0
    #     last_ntimes_memory = 0
    #     last_device = None
    #     cur_cpu_peak = None
    #     cur_memory_peak = None
    #     
    #     # Loop over the raw sheet to fill capacity values database
    #     for _, row in df_input1.iterrows():
    #         device_name = row["device_name"]
    #         
    #         # Skip empty rows or rows with device_name == "1"
    #         if row.isnull().all() or device_name == "1":
    #             continue
    #         
    #         ntimes_cpu = row.iloc[3]
    #         ntimes_memory = row.iloc[6]
    #         
    #         # Update last device tracking
    #         if pd.notna(device_name):
    #             last_ntimes_cpu = ntimes_cpu
    #             last_ntimes_memory = ntimes_memory
    #             last_device = device_name
    #             cur_cpu_peak = None
    #             cur_memory_peak = None
    #         
    #         # Handle peak values when ntimes are 0
    #         if ntimes_cpu == 0 and ntimes_memory == 0:
    #             cur_cpu_peak = convert_NaN_to_None(row.get("peak_cpu"))
    #             cur_memory_peak = convert_NaN_to_None(row.get("peak_memory"))
    #         elif ntimes_memory == 0:
    #             cur_memory_peak = convert_NaN_to_None(row.get("peak_memory"))
    #         elif ntimes_cpu == 0:
    #             cur_cpu_peak = convert_NaN_to_None(row.get("peak_cpu"))
    #         
    #         # Insert initial capacity values if both ntimes are 0
    #         if ntimes_cpu == 0 and ntimes_memory == 0:
    #             insert_capacity_values(
    #                 db,
    #                 device_name,
    #                 changeFormat(convert_NaN_to_None(row.get("mean_cpu"))),
    #                 changeFormat(cur_cpu_peak),
    #                 0,
    #                 changeFormat(convert_NaN_to_None(row.get("mean_memory"))),
    #                 changeFormat(cur_memory_peak),
    #                 0
    #             )
    #             continue
    #         
    #         # Current ntimes values
    #         cur_ntimes_cpu = last_ntimes_cpu
    #         cur_ntimes_memory = last_ntimes_memory
    #         
    #         # Process CPU details
    #         if cur_ntimes_cpu > 0:
    #             cpu_details = extract_date_time_peak(row.get("first_peak_cpu_dt", ""))
    #             if (cpu_details
    #                 and cpu_details[2] is not None
    #                 and is_between(cpu_details[0], start_date, end_date)):
    #             
    #                 update_ntime_cpu(db, last_device)
    #                 
    #                 # Update CPU peak if higher
    #                 if cur_cpu_peak is None or float(cpu_details[2]) > cur_cpu_peak:
    #                     cur_cpu_peak = float(cpu_details[2])
    #                     cpu_date = cpu_details[0]
    #                     cpu_time = cpu_details[1]
    #                     cpu_alert_duration = (
    #                         datetime.strptime(cpu_details[3], "%H:%M")
    #                         - datetime.strptime(cpu_details[1], "%H:%M")
    #                     ).total_seconds() / 60
    #                 
    #                     update_capacity_peaks_cpu(
    #                         db,
    #                         last_device,
    #                         cur_cpu_peak,
    #                         cpu_date,
    #                         cpu_time,
    #                         cpu_alert_duration
    #                     )
    #         
    #         # Process Memory details
    #         if cur_ntimes_memory > 0:
    #             memory_details = extract_date_time_peak(row.get("first_peak_memory_dt", ""))
    #             if (memory_details
    #                 and memory_details[2] is not None
    #                 and is_between(memory_details[0], start_date, end_date)):
    #             
    #                 update_ntime_memory(db, last_device)
    #                 
    #                 # Update Memory peak if higher
    #                 if cur_memory_peak is None or float(memory_details[2]) > cur_memory_peak:
    #                     cur_memory_peak = float(memory_details[2])
    #                     memory_date = memory_details[0]
    #                     memory_time = memory_details[1]
    #                     memory_alert_duration = (
    #                         datetime.strptime(memory_details[3], "%H:%M")
    #                         - datetime.strptime(memory_details[1], "%H:%M")
    #                     ).total_seconds() / 60
    #                 
    #                     update_capacity_peaks_memory(
    #                         db,
    #                         last_device,
    #                         cur_memory_peak,
    #                         memory_date,
    #                         memory_time,
    #                         memory_alert_duration
    #                     )
    #         
    #         # Update last ntimes values
    #         last_ntimes_cpu = cur_ntimes_cpu
    #         last_ntimes_memory = cur_ntimes_memory
    #     
    #     # Process connection count file (df_input2)
    #     logger.info("Processing connection count data")
    #     for _, row in df_input2.iterrows():
    #         if pd.isna(row["Entity Name"]) or row["Entity Name"] == "1":
    #             continue
    #         
    #         device_name = row["Entity Name"]
    #         
    #         update_connection_in_capacity_values(
    #             db, 
    #             device_name, 
    #             changeFormat(convert_NaN_to_None(row["Mean"])), 
    #             changeFormat(convert_NaN_to_None(row["Peak"]))
    #         )
    # 
    # except Exception as e:
    #     logger.error(f"Error processing capacity files: {str(e)}", exc_info=True)
    #     db.rollback()
    #     raise HTTPException(
    #         status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
    #         detail=f"Failed to process capacity files: {str(e)}"
    #     )

@router.post("/upload")
async def upload_capacity_report(
    start_date: str = Form(...),
    end_date: str = Form(...),
    raw_data_file: UploadFile = File(...),
    connection_count_file: UploadFile = File(...),
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    """
    Accept capacity dashboard inputs: start/end dates and two files.
    Files are saved to persistent storage for later processing.
    """
    try:
        # Create timestamp-based directory for this upload
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        upload_dir = CAPACITY_UPLOADS_DIR / timestamp
        upload_dir.mkdir(parents=True, exist_ok=True)
        
        # Save raw data file
        raw_data_path = upload_dir / f"raw_data_{raw_data_file.filename or 'file.xlsx'}"
        raw_data_content = await raw_data_file.read()
        with open(raw_data_path, "wb") as f:
            f.write(raw_data_content)
        logger.info(f"Saved raw data file to: {raw_data_path}")
        
        # Save connection count file
        connection_count_path = upload_dir / f"connection_count_{connection_count_file.filename or 'file.xlsx'}"
        connection_count_content = await connection_count_file.read()
        with open(connection_count_path, "wb") as f:
            f.write(connection_count_content)
        logger.info(f"Saved connection count file to: {connection_count_path}")
        
        # Process capacity files using stored file paths
        await process_capacity_files(
            db=db,
            start_date=start_date,
            end_date=end_date,
            raw_data_file_path=str(raw_data_path),
            connection_count_file_path=str(connection_count_path),
        )

        return {
            "message": "Capacity report files uploaded and saved successfully.",
            "start_date": start_date,
            "end_date": end_date,
            "raw_data_filename": raw_data_file.filename,
            "connection_count_filename": connection_count_file.filename,
            "upload_directory": str(upload_dir),
            "raw_data_path": str(raw_data_path),
            "connection_count_path": str(connection_count_path),
        }
    except Exception as e:
        logger.error(f"Error uploading capacity files: {str(e)}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to upload capacity files: {str(e)}"
        )


@router.get("/export-summary")
async def export_capacity_summary(
    start_date: Optional[str] = Query(None, description="Optional: Start date filter (format: DD-MMM-YYYY, e.g., 01-Feb-2025)"),
    end_date: Optional[str] = Query(None, description="Optional: End date filter (format: DD-MMM-YYYY, e.g., 28-Feb-2025)"),
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    """
    Generate capacity summary Excel with separate sheets per region
    and production/non-production hours.
    
    Optionally filters by start_date and end_date using cpu_date and memory_date columns.
    For CPU categories, filters by cpu_date. For Memory categories, filters by memory_date.

    Sheet layout (per region):
        Zone Name
        Total Device Count
        CPU Green
        Memory Green
        CPU Yellow
        Memory Yellow
        CPU Red
        Memory Red
    """

    def get_zones_for_region(region_prefix: str) -> List[str]:
        """Return distinct zone names for a region prefix (e.g. 'XYZ')."""
        results: List[Tuple[str]] = (
            db.query(ZoneDeviceMapping.zone_name)
            .filter(ZoneDeviceMapping.zone_name.startswith(region_prefix))
            .distinct()
            .order_by(ZoneDeviceMapping.zone_name)
            .all()
        )
        return [row[0] for row in results]

    def compute_category(
        zone_name: str,
        prod_hours: bool,
        peak_min: int,
        peak_max: int,
        attribute,
        time_column,
        date_column,
    ) -> int:
        """
        Count distinct devices for a zone that fall into a threshold bucket
        for the given attribute (CPU/Memory) and time-of-day window.
        Optionally filters by date_column if start_date/end_date are provided.
        """
        # Production: 09:00-16:00, Non-production: everything else
        time_filter = (
            time_column.between("09:00", "16:00")
            if prod_hours
            else ~time_column.between("09:00", "16:00")
        )

        query = (
            db.query(func.count(CapacityValues.device_name.distinct()))
            .join(
                ZoneDeviceMapping,
                CapacityValues.device_name == ZoneDeviceMapping.device_name,
            )
            .filter(
                ZoneDeviceMapping.zone_name == zone_name,
                attribute.between(peak_min, peak_max),
                time_filter,
            )
        )
        
        # Apply date filters if provided
        if start_date or end_date:
            engine = get_engine()
            if engine and engine.dialect.name == 'oracle':
                # Use Oracle TO_DATE for proper date comparison
                # Only filter if date_column is not null
                date_conditions = [date_column.isnot(None)]
                if start_date:
                    date_conditions.append(
                        text(f"TO_DATE({date_column.key}, 'DD-MON-YYYY') >= TO_DATE(:start_date_param, 'DD-MON-YYYY')")
                    )
                if end_date:
                    date_conditions.append(
                        text(f"TO_DATE({date_column.key}, 'DD-MON-YYYY') <= TO_DATE(:end_date_param, 'DD-MON-YYYY')")
                    )
                query = query.filter(and_(*date_conditions))
                # Bind parameters
                if start_date:
                    query = query.params(start_date_param=start_date)
                if end_date:
                    query = query.params(end_date_param=end_date)
            else:
                # Fallback to string comparison for other databases
                date_conditions = [date_column.isnot(None)]
                if start_date:
                    date_conditions.append(date_column >= start_date)
                if end_date:
                    date_conditions.append(date_column <= end_date)
                query = query.filter(and_(*date_conditions))

        count = query.scalar()
        return int(count or 0)

    def build_region_summary(region_prefix: str, prod_hours: bool) -> List[List]:
        """
        Build summary rows for a region and production flag.

        Returns list of:
            [zone, total_devices, cpu_green, memory_green,
             cpu_yellow, memory_yellow, cpu_red, memory_red]
        """
        zones = get_zones_for_region(region_prefix)
        summary_rows: List[List] = []

        for zone in zones:
            # Total devices (no time filter)
            total_devices = (
                db.query(func.count(ZoneDeviceMapping.device_name.distinct()))
                .filter(ZoneDeviceMapping.zone_name == zone)
                .scalar()
            )
            total_devices = int(total_devices or 0)

            # CPU categories - filter by cpu_date
            cpu_red = compute_category(
                zone,
                prod_hours,
                71,
                100,
                CapacityValues.peak_cpu,
                CapacityValues.cpu_time,
                CapacityValues.cpu_date,
            )
            cpu_yellow = compute_category(
                zone,
                prod_hours,
                61,
                70,
                CapacityValues.peak_cpu,
                CapacityValues.cpu_time,
                CapacityValues.cpu_date,
            )
            cpu_green = max(total_devices - (cpu_red + cpu_yellow), 0)

            # Memory categories - filter by memory_date
            memory_red = compute_category(
                zone,
                prod_hours,
                71,
                100,
                CapacityValues.peak_memory,
                CapacityValues.memory_time,
                CapacityValues.memory_date,
            )
            memory_yellow = compute_category(
                zone,
                prod_hours,
                61,
                70,
                CapacityValues.peak_memory,
                CapacityValues.memory_time,
                CapacityValues.memory_date,
            )
            memory_green = max(total_devices - (memory_red + memory_yellow), 0)

            summary_rows.append(
                [
                    zone,
                    total_devices,
                    cpu_green,
                    memory_green,
                    cpu_yellow,
                    memory_yellow,
                    cpu_red,
                    memory_red,
                ]
            )

        # Sort zones alphabetically
        summary_rows.sort(key=lambda r: r[0] or "")
        return summary_rows

    headers = [
        "Zone Name",
        "Total Device Count",
        "CPU Green",
        "Memory Green",
        "CPU Yellow",
        "Memory Yellow",
        "CPU Red",
        "Memory Red",
    ]

    # Create workbook with separate sheets for each region and prod flag
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    regions = ["XYZ", "ORM-XYZ", "DRM", "ORM-DRM"]

    for region in regions:
        # Production sheet
        prod_rows = build_region_summary(region, prod_hours=True)
        if prod_rows:
            ws_prod = wb.create_sheet(title=f"Production {region}")
            ws_prod.append(headers)
            for row in prod_rows:
                ws_prod.append(row)

            for col_cells in ws_prod.columns:
                length = max(
                    len(str(cell.value)) if cell.value is not None else 0
                    for cell in col_cells
                )
                ws_prod.column_dimensions[col_cells[0].column_letter].width = min(
                    max(length + 2, 12), 40
                )

        # Non-production sheet
        non_prod_rows = build_region_summary(region, prod_hours=False)
        if non_prod_rows:
            ws_non = wb.create_sheet(title=f"Non-Production {region}")
            ws_non.append(headers)
            for row in non_prod_rows:
                ws_non.append(row)

            for col_cells in ws_non.columns:
                length = max(
                    len(str(cell.value)) if cell.value is not None else 0
                    for cell in col_cells
                )
                ws_non.column_dimensions[col_cells[0].column_letter].width = min(
                    max(length + 2, 12), 40
                )

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename=\"capacity_summary.xlsx\"'},
    )


@router.get("/export")
async def export_capacity_details(
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    """
    Export per-device capacity values into separate region sheets,
    with one row per device (similar to the raw capacity report view).
    """
    # Build device -> zone map (take the first mapping per device)
    zone_map: Dict[str, str] = {}
    mappings = db.query(ZoneDeviceMapping).all()
    for m in mappings:
        if m.device_name and m.device_name not in zone_map:
            zone_map[m.device_name] = m.zone_name

    rows = (
        db.query(
            CapacityValues.device_name,
            CapacityValues.mean_cpu,
            CapacityValues.peak_cpu,
            CapacityValues.mean_memory,
            CapacityValues.peak_memory,
            CapacityValues.mean_connection,
            CapacityValues.peak_connection,
        )
        .order_by(CapacityValues.device_name)
        .all()
    )

    # Group rows by region prefix
    region_data: Dict[str, list] = {
        "XYZ": [],
        "ORM-XYZ": [],
        "DRM": [],
        "ORM-DRM": [],
    }

    headers = [
        "Zone Name",
        "Device Name",
        "Mean CPU",
        "Peak CPU",
        "Mean Memory",
        "Peak Memory",
        "Mean Connection",
        "Peak Connection",
    ]

    for r in rows:
        device = r[0] or ""
        zone_name = zone_map.get(device, "")

        # Determine region based on zone name prefix
        region = None
        if zone_name.startswith("ORM-XYZ"):
            region = "ORM-XYZ"
        elif zone_name.startswith("XYZ"):
            region = "XYZ"
        elif zone_name.startswith("ORM-DRM"):
            region = "ORM-DRM"
        elif zone_name.startswith("DRM"):
            region = "DRM"

        if region:
            region_data[region].append(
                [
                    zone_name,
                    device,
                    r[1],
                    r[2],
                    r[3],
                    r[4],
                    r[5],
                    r[6],
                ]
            )

    # Create workbook with separate sheets for each region
    wb = openpyxl.Workbook()
    
    sheets_created = False
    # Create sheets for each region (only if they have data)
    for region_name in ["XYZ", "ORM-XYZ", "DRM", "ORM-DRM"]:
        if region_data[region_name]:
            # Sort by zone name first, then by device name
            region_data[region_name].sort(key=lambda x: (x[0] or "", x[1] or ""))

            ws = wb.create_sheet(title=region_name)
            ws.append(headers)

            for row_data in region_data[region_name]:
                ws.append(row_data)

            # Autosize columns for this sheet
            for col_cells in ws.columns:
                length = max(
                    len(str(cell.value)) if cell.value is not None else 0
                    for cell in col_cells
                )
                ws.column_dimensions[col_cells[0].column_letter].width = min(
                    max(length + 2, 12), 40
                )
            sheets_created = True

    if sheets_created:
        # Remove the default "Sheet" if we created specific sheets
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
    else:
        # No data found, rename default sheet
        ws = wb.active
        ws.title = "No Data"
        ws.append(["No devices found"])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename=\"capacity_devices.xlsx\"'},
    )


@router.get("/dashboard")
async def get_capacity_dashboard(
    region: str = Query(..., description="Region name: XYZ, ORM-XYZ, DRM, or ORM-DRM"),
    production_hours: bool = Query(True, description="True for production hours (09:00-16:00), False for non-production"),
    zone_name: Optional[str] = Query(None, description="Optional: Get device details for a specific zone"),
    start_date: Optional[str] = Query(None, description="Optional: Start date filter (format: DD-MMM-YYYY, e.g., 01-Feb-2025)"),
    end_date: Optional[str] = Query(None, description="Optional: End date filter (format: DD-MMM-YYYY, e.g., 28-Feb-2025)"),
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    """
    Get capacity dashboard data for a region.
    Returns zone summary and optionally device details for a specific zone.
    Optionally filters by start_date and end_date using cpu_date and memory_date columns.
    """
    # Validate region
    valid_regions = ["XYZ", "ORM-XYZ", "DRM", "ORM-DRM"]
    if region not in valid_regions:
        from fastapi import HTTPException, status
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Invalid region. Must be one of: {', '.join(valid_regions)}"
        )

    # Get zones for the region from region_zone_mapping
    zones_query = (
        db.query(RegionZoneMapping.zone_name)
        .filter(RegionZoneMapping.region_name == region)
        .order_by(RegionZoneMapping.zone_name)
        .all()
    )
    zones = [z[0] for z in zones_query]

    def compute_category_count(
        zone_name: str,
        prod_hours: bool,
        peak_min: float,
        peak_max: float,
        attribute,
        time_column,
        date_column,
    ) -> int:
        """Count devices in a category (Normal/Warning/Critical) for a zone."""
        # Production: 09:00-16:00, Non-production: everything else
        if prod_hours:
            time_filter = time_column.between("09:00", "16:00")
        else:
            time_filter = or_(
                time_column < "09:00",
                time_column > "16:00",
                time_column.is_(None)
            )

        query = (
            db.query(CapacityValues.device_name)
            .join(
                ZoneDeviceMapping,
                CapacityValues.device_name == ZoneDeviceMapping.device_name,
            )
            .filter(
                ZoneDeviceMapping.zone_name == zone_name,
                attribute.between(peak_min, peak_max),
                time_filter,
            )
        )
        
        # Apply date filters if provided
        if start_date or end_date:
            engine = get_engine()
            if engine and engine.dialect.name == 'oracle':
                # Use Oracle TO_DATE for proper date comparison
                # Only filter if date_column is not null
                date_conditions = [date_column.isnot(None)]
                if start_date:
                    date_conditions.append(
                        text(f"TO_DATE({date_column.key}, 'DD-MON-YYYY') >= TO_DATE(:start_date_param, 'DD-MON-YYYY')")
                    )
                if end_date:
                    date_conditions.append(
                        text(f"TO_DATE({date_column.key}, 'DD-MON-YYYY') <= TO_DATE(:end_date_param, 'DD-MON-YYYY')")
                    )
                query = query.filter(and_(*date_conditions))
                # Bind parameters using bindparam
                if start_date:
                    query = query.params(start_date_param=start_date)
                if end_date:
                    query = query.params(end_date_param=end_date)
            else:
                # Fallback to string comparison for other databases
                date_conditions = [date_column.isnot(None)]
                if start_date:
                    date_conditions.append(date_column >= start_date)
                if end_date:
                    date_conditions.append(date_column <= end_date)
                query = query.filter(and_(*date_conditions))
        
        count = query.distinct().count()
        return int(count or 0)

    # Build zone summary
    zone_summary = []
    for zone in zones:
        # Total device count (no time filter)
        total_devices = (
            db.query(ZoneDeviceMapping.device_name)
            .filter(ZoneDeviceMapping.zone_name == zone)
            .distinct()
            .count()
        )
        total_devices = int(total_devices or 0)

        # CPU categories
        cpu_critical = compute_category_count(zone, production_hours, 71, 100, CapacityValues.peak_cpu, CapacityValues.cpu_time, CapacityValues.cpu_date)
        cpu_warning = compute_category_count(zone, production_hours, 61, 70, CapacityValues.peak_cpu, CapacityValues.cpu_time, CapacityValues.cpu_date)
        cpu_normal = max(total_devices - (cpu_critical + cpu_warning), 0)

        # Memory categories
        memory_critical = compute_category_count(zone, production_hours, 71, 100, CapacityValues.peak_memory, CapacityValues.memory_time, CapacityValues.memory_date)
        memory_warning = compute_category_count(zone, production_hours, 61, 70, CapacityValues.peak_memory, CapacityValues.memory_time, CapacityValues.memory_date)
        memory_normal = max(total_devices - (memory_critical + memory_warning), 0)

        zone_summary.append({
            "zone_name": zone,
            "total_device_count": total_devices,
            "cpu_normal": cpu_normal,
            "cpu_warning": cpu_warning,
            "cpu_critical": cpu_critical,
            "memory_normal": memory_normal,
            "memory_warning": memory_warning,
            "memory_critical": memory_critical,
        })

    response = {
        "region": region,
        "production_hours": production_hours,
        "zone_summary": zone_summary,
    }
    
    # Add date filters to response if provided
    if start_date:
        response["start_date"] = start_date
    if end_date:
        response["end_date"] = end_date

    # If zone_name is provided, also return device details for that zone
    if zone_name:
        # Verify zone belongs to the region
        zone_exists = (
            db.query(RegionZoneMapping)
            .filter(
                RegionZoneMapping.region_name == region,
                RegionZoneMapping.zone_name == zone_name
            )
            .first()
        )
        
        if not zone_exists:
            from fastapi import HTTPException, status
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Zone '{zone_name}' not found in region '{region}'"
            )

        # Get devices for this zone
        devices_query = (
            db.query(
                CapacityValues.device_name,
                CapacityValues.peak_cpu,
                CapacityValues.ntimes_cpu,
                CapacityValues.cpu_alert_duration,
                CapacityValues.peak_memory,
                CapacityValues.ntimes_memory,
                CapacityValues.memory_alert_duration,
            )
            .join(
                ZoneDeviceMapping,
                CapacityValues.device_name == ZoneDeviceMapping.device_name,
            )
            .filter(ZoneDeviceMapping.zone_name == zone_name)
        )
        
        # Apply date filters if provided
        # For device details, we include devices if either cpu_date or memory_date falls within the range
        if start_date or end_date:
            engine = get_engine()
            if engine and engine.dialect.name == 'oracle':
                # Use Oracle TO_DATE - include device if either cpu_date or memory_date is in range
                date_conditions = []
                if start_date:
                    date_conditions.append(
                        or_(
                            and_(
                                CapacityValues.cpu_date.isnot(None),
                                text("TO_DATE(cpu_date, 'DD-MON-YYYY') >= TO_DATE(:start_date_param, 'DD-MON-YYYY')")
                            ),
                            and_(
                                CapacityValues.memory_date.isnot(None),
                                text("TO_DATE(memory_date, 'DD-MON-YYYY') >= TO_DATE(:start_date_param, 'DD-MON-YYYY')")
                            ),
                            and_(
                                CapacityValues.cpu_date.is_(None),
                                CapacityValues.memory_date.is_(None)
                            )
                        )
                    )
                if end_date:
                    date_conditions.append(
                        or_(
                            and_(
                                CapacityValues.cpu_date.isnot(None),
                                text("TO_DATE(cpu_date, 'DD-MON-YYYY') <= TO_DATE(:end_date_param, 'DD-MON-YYYY')")
                            ),
                            and_(
                                CapacityValues.memory_date.isnot(None),
                                text("TO_DATE(memory_date, 'DD-MON-YYYY') <= TO_DATE(:end_date_param, 'DD-MON-YYYY')")
                            ),
                            and_(
                                CapacityValues.cpu_date.is_(None),
                                CapacityValues.memory_date.is_(None)
                            )
                        )
                    )
                for condition in date_conditions:
                    devices_query = devices_query.filter(condition)
                if start_date:
                    devices_query = devices_query.params(start_date_param=start_date)
                if end_date:
                    devices_query = devices_query.params(end_date_param=end_date)
            else:
                # Fallback string comparison
                if start_date:
                    devices_query = devices_query.filter(
                        or_(
                            and_(CapacityValues.cpu_date.isnot(None), CapacityValues.cpu_date >= start_date),
                            and_(CapacityValues.memory_date.isnot(None), CapacityValues.memory_date >= start_date),
                            and_(CapacityValues.cpu_date.is_(None), CapacityValues.memory_date.is_(None))
                        )
                    )
                if end_date:
                    devices_query = devices_query.filter(
                        or_(
                            and_(CapacityValues.cpu_date.isnot(None), CapacityValues.cpu_date <= end_date),
                            and_(CapacityValues.memory_date.isnot(None), CapacityValues.memory_date <= end_date),
                            and_(CapacityValues.cpu_date.is_(None), CapacityValues.memory_date.is_(None))
                        )
                    )
        
        devices_query = devices_query.distinct().order_by(CapacityValues.device_name).all()

        device_details = []
        for device in devices_query:
            device_details.append({
                "device_name": device[0] or "",
                "cpu_peak_percent": float(device[1]) if device[1] is not None else 0.0,
                "cpu_peak_count": int(device[2]) if device[2] is not None else 0,
                "cpu_peak_duration_min": float(device[3]) if device[3] is not None else 0.0,
                "memory_peak_percent": float(device[4]) if device[4] is not None else 0.0,
                "memory_peak_count": int(device[5]) if device[5] is not None else 0,
                "memory_peak_duration_min": float(device[6]) if device[6] is not None else 0.0,
            })

        response["device_details"] = device_details

    return response


# Pydantic models for device/zone management
class DeviceZoneMappingRequest(BaseModel):
    zone_name: str
    device_name: str


class ZoneRegionMappingRequest(BaseModel):
    region_name: str
    zone_name: str
    action: Optional[str] = None  # Not needed for separate routes, kept for backward compatibility
    new_zone_name: Optional[str] = None  # For editing zone name


class EditZoneRequest(BaseModel):
    old_zone_name: str
    new_zone_name: str
    region_name: str


@router.get("/zones")
async def get_all_zones(
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    """Get all zones from region_zone_mapping"""
    zones = (
        db.query(RegionZoneMapping.zone_name, RegionZoneMapping.region_name)
        .order_by(RegionZoneMapping.region_name, RegionZoneMapping.zone_name)
        .all()
    )
    return {
        "zones": [{"zone_name": z[0], "region_name": z[1]} for z in zones]
    }


@router.get("/devices")
async def get_all_devices(
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    """Get all unique device names from capacity_values"""
    devices = (
        db.query(CapacityValues.device_name)
        .filter(CapacityValues.device_name.isnot(None))
        .distinct()
        .order_by(CapacityValues.device_name)
        .all()
    )
    return {
        "devices": [d[0] for d in devices if d[0]]
    }


@router.get("/regions")
async def get_all_regions(
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    """Get all unique regions from region_zone_mapping"""
    regions = (
        db.query(RegionZoneMapping.region_name)
        .distinct()
        .order_by(RegionZoneMapping.region_name)
        .all()
    )
    return {
        "regions": [r[0] for r in regions]
    }


@router.get("/zone-device-mappings")
async def get_zone_device_mappings(
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    """Get all zone-device mappings"""
    mappings = (
        db.query(ZoneDeviceMapping.zone_name, ZoneDeviceMapping.device_name)
        .order_by(ZoneDeviceMapping.zone_name, ZoneDeviceMapping.device_name)
        .all()
    )
    return {
        "mappings": [{"zone_name": m[0], "device_name": m[1]} for m in mappings]
    }

@router.post("/device-zone-mapping/add", status_code=status.HTTP_201_CREATED)
async def add_device_to_zone(
    request: DeviceZoneMappingRequest,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    """Add a device to a zone (create zone-device mapping)"""
    logger.info(f"=== add_device_to_zone FUNCTION CALLED ===")
    logger.info(f"add_device_to_zone called with zone_name='{request.zone_name}', device_name='{request.device_name}'")
    # Check if device is already present in that zone in zone_device_mapping
    logger.info(f"Checking if device '{request.device_name}' exists in zone '{request.zone_name}'")
    existing = (
        db.query(ZoneDeviceMapping)
        .filter(
            ZoneDeviceMapping.zone_name == request.zone_name,
            ZoneDeviceMapping.device_name == request.device_name
        )
        .first()
    )
    logger.info(f"Existing mapping check result: {existing is not None}")
    if existing:
        logger.warning(f"Device '{request.device_name}' already mapped to zone '{request.zone_name}'")
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"The device name '{request.device_name}' you are trying to add is already present in zone '{request.zone_name}'"
        )
    
    logger.info(f"Device '{request.device_name}' not found in zone '{request.zone_name}', proceeding to add")
    # Create device-zone mapping using ORM
    
    logger.info(f"Creating mapping for device '{request.device_name}' in zone '{request.zone_name}'")
    try:
        new_mapping = ZoneDeviceMapping(
            zone_name=request.zone_name,
            device_name=request.device_name
        )
        db.add(new_mapping)
        db.commit()
        
        return {
            "message": f"Device '{request.device_name}' successfully added to zone '{request.zone_name}'",
            "zone_name": request.zone_name,
            "device_name": request.device_name
        }
    except Exception as e:
        db.rollback()
        logger.error(f"Exception while adding device to zone: {str(e)}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to add device to zone: {str(e)}"
        )


@router.put("/device-zone-mapping/update")
async def update_device_zone_mapping(
    old_zone_name: str = Query(...),
    old_device_name: str = Query(...),
    new_zone_name: str = Query(None),
    new_device_name: str = Query(None),
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    """Update a device-zone mapping (move device to different zone or rename device)"""
    # Find existing mapping
    existing = (
        db.query(ZoneDeviceMapping)
        .filter(
            ZoneDeviceMapping.zone_name == old_zone_name,
            ZoneDeviceMapping.device_name == old_device_name
        )
        .first()
    )
    if not existing:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Mapping not found: device '{old_device_name}' in zone '{old_zone_name}'"
        )

    # Update zone if provided
    if new_zone_name and new_zone_name != old_zone_name:
        # Verify new zone exists
        zone_exists = (
            db.query(RegionZoneMapping)
            .filter(RegionZoneMapping.zone_name == new_zone_name)
            .first()
        )
        if not zone_exists:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Zone '{new_zone_name}' not found"
            )
        existing.zone_name = new_zone_name

    # Update device if provided
    if new_device_name and new_device_name != old_device_name:
        # Simply update the device name in the mapping - no need to check if it exists in capacity_values
        existing.device_name = new_device_name

    try:
        db.commit()
        return {
            "message": "Device-zone mapping updated successfully",
            "zone_name": existing.zone_name,
            "device_name": existing.device_name
        }
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to update mapping: {str(e)}"
        )


@router.delete("/device-zone-mapping/delete")
async def delete_device_zone_mapping(
    zone_name: str = Query(...),
    device_name: str = Query(...),
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    """Delete a device-zone mapping"""
    mapping = (
        db.query(ZoneDeviceMapping)
        .filter(
            ZoneDeviceMapping.zone_name == zone_name,
            ZoneDeviceMapping.device_name == device_name
        )
        .first()
    )
    if not mapping:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Mapping not found: device '{device_name}' in zone '{zone_name}'"
        )

    try:
        db.delete(mapping)
        db.commit()
        return {
            "message": f"Device '{device_name}' removed from zone '{zone_name}'"
        }
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to delete mapping: {str(e)}"
        )


@router.post("/zone-region-mapping/add", status_code=status.HTTP_201_CREATED)
async def add_zone_to_region(
    request: ZoneRegionMappingRequest,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    """Add a zone to a region"""
    # Check if zone already exists
    existing = (
        db.query(RegionZoneMapping)
        .filter(RegionZoneMapping.zone_name == request.zone_name)
        .first()
    )
    if existing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Zone '{request.zone_name}' already exists"
        )

    # Create zone-region mapping using ORM
    try:
        new_mapping = RegionZoneMapping(
            region_name=request.region_name,
            zone_name=request.zone_name
        )
        db.add(new_mapping)
        db.commit()
        
        return {
            "message": f"Zone '{request.zone_name}' added to region '{request.region_name}'",
            "region_name": request.region_name,
            "zone_name": request.zone_name
        }
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to add zone: {str(e)}"
        )


@router.put("/zone-region-mapping/update")
async def update_zone_region_mapping(
    request: ZoneRegionMappingRequest,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    """Update a zone in a region"""
    # Find existing zone
    existing = (
        db.query(RegionZoneMapping)
        .filter(RegionZoneMapping.zone_name == request.zone_name)
        .first()
    )
    if not existing:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Zone '{request.zone_name}' not found"
        )

    # Update region if changed
    if request.region_name != existing.region_name:
        existing.region_name = request.region_name

    # Update zone name if new_zone_name is provided
    if request.new_zone_name and request.new_zone_name != request.zone_name:
        # Check if new zone name already exists
        zone_exists = (
            db.query(RegionZoneMapping)
            .filter(RegionZoneMapping.zone_name == request.new_zone_name)
            .first()
        )
        if zone_exists:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Zone '{request.new_zone_name}' already exists"
            )

        # Update all zone_device_mapping entries
        device_mappings = (
            db.query(ZoneDeviceMapping)
            .filter(ZoneDeviceMapping.zone_name == request.zone_name)
            .all()
        )
        for dm in device_mappings:
            dm.zone_name = request.new_zone_name

        existing.zone_name = request.new_zone_name

    try:
        db.commit()
        return {
            "message": f"Zone updated successfully",
            "region_name": existing.region_name,
            "zone_name": existing.zone_name
        }
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to update zone: {str(e)}"
        )


@router.delete("/zone-region-mapping/delete")
async def delete_zone_region_mapping(
    zone_name: str = Query(...),
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    """Delete a zone from a region"""
    # Find existing zone
    existing = (
        db.query(RegionZoneMapping)
        .filter(RegionZoneMapping.zone_name == zone_name)
        .first()
    )
    if not existing:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Zone '{zone_name}' not found"
        )

    # Also delete all device mappings for this zone
    device_mappings = (
        db.query(ZoneDeviceMapping)
        .filter(ZoneDeviceMapping.zone_name == zone_name)
        .all()
    )
    for dm in device_mappings:
        db.delete(dm)

    try:
        db.delete(existing)
        db.commit()
        return {
            "message": f"Zone '{zone_name}' and all its device mappings deleted"
        }
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to delete zone: {str(e)}"
        )
