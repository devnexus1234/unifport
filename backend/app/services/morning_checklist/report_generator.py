from datetime import date, timedelta
from typing import Optional
from io import BytesIO
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from sqlalchemy.orm import Session
from app.models.morning_checklist import MorningChecklist

def _get_prev_date(target_date: date) -> date:
    return target_date - timedelta(days=1)

def _load_rows(
    db: Session,
    target_date: date,
    application_name: Optional[str] = None,
    asset_owner: Optional[str] = None,
):
    query = db.query(MorningChecklist).filter(MorningChecklist.mc_check_date == target_date)
    if application_name:
        query = query.filter(MorningChecklist.application_name == application_name)
    if asset_owner:
        query = query.filter(MorningChecklist.asset_owner == asset_owner)
    return query.all()

def generate_morning_checklist_excel(
    db: Session,
    target_date: date,
    application_name: Optional[str] = None,
    asset_owner: Optional[str] = None
) -> BytesIO:
    """
    Generates an Excel report for the morning checklist.
    Returns a BytesIO object containing the Excel file.
    """
    
    # Fetch data
    current_rows = _load_rows(db, target_date, application_name, asset_owner)
    prev_date = _get_prev_date(target_date)
    previous_rows = _load_rows(db, prev_date, application_name, asset_owner)

    # Build maps: hostname -> list of rows
    prev_map = {}
    for r in previous_rows:
        key = (r.hostname, r.commands or "")
        if key not in prev_map:
            prev_map[key] = r

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Checklist {target_date}"
    
    headers = [
        "Hostname", 
        "IP", 
        "Application", 
        "Asset Owner", 
        "Command", 
        f"Output ({prev_date})", 
        f"Output ({target_date})", 
        "Diff Status"
    ]
    ws.append(headers)
    
    # Styling
    fill_diff = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    font_bold = Font(bold=True)
    alignment_wrap = Alignment(wrap_text=True)
    
    # Apply header style
    for cell in ws[1]:
        cell.font = font_bold

    for row in current_rows:
        hostname = row.hostname
        command = row.commands or ""
        
        # Find prev
        prev_row = prev_map.get((hostname, command))
        
        output_curr = row.mc_output or ""
        output_prev = prev_row.mc_output if prev_row else ""
        
        # Determine strict diff
        has_diff = output_curr.strip() != output_prev.strip()
        diff_status = "Diff" if has_diff else "No Diff"
        if not prev_row:
             diff_status = "New / No History"

        ws.append([
            row.hostname,
            row.ip,
            row.application_name,
            row.asset_owner,
            command,
            output_prev,
            output_curr,
            diff_status
        ])
        
        if has_diff:
            # Highlight the row
            current_row_idx = ws.max_row
            for cell in ws[current_row_idx]:
                cell.fill = fill_diff

    # Auto-adjust column width (approximate)
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            cell.alignment = alignment_wrap # Wrap text for all cells
            try:
                val_len = len(str(cell.value))
                if val_len > max_length:
                    max_length = val_len
            except:
                pass
        # Cap width
        adjusted_width = min(max_length + 2, 80)
        ws.column_dimensions[column].width = adjusted_width
    
    # Set explicit widths for output columns
    ws.column_dimensions['F'].width = 50
    ws.column_dimensions['G'].width = 50
    ws.column_dimensions['E'].width = 40

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    return stream
